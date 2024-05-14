import pytest
import openpyxl
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from constants import sauceDemoConsts
from time import sleep

class Test_SauceDemo:
    def setup_method(self):
        self.driver = webdriver.Chrome()
        # self.driver.maximize_window()
        self.driver.get(sauceDemoConsts.BASE_URL)

    def teardown_method(self):
        self.driver.quit()

    def login(self, username, password):
        username_input = self.driver.find_element(By.ID, sauceDemoConsts.username_id)
        password_input = self.driver.find_element(By.ID, sauceDemoConsts.password_id)
        login_button = self.driver.find_element(By.ID, sauceDemoConsts.login_button_id)
        
        username_input.clear()
        username_input.send_keys(username)
        password_input.clear()
        password_input.send_keys(password)
        login_button.click()

    #Eski ödevdeki;
        # -Kullanıcı adı ve şifre alanları boş geçildiğinde uyarı mesajı olarak "Epic sadface: Username is required" gösterilmelidir.
        # -Sadece şifre alanı boş geçildiğinde uyarı mesajı olarak "Epic sadface: Password is required" gösterilmelidir.
        # -Kullanıcı adı "locked_out_user" şifre alanı "secret_sauce" gönderildiğinde "Epic sadface: Sorry, this user has been locked out." mesajı gösterilmelidir.
    #kısımlarını tek fonksiyonda birleştirip parametrize'ı excel ile birlikte buradaki hatalı girişler için kullanmak istedim.
    def getExcelData():
        excelFile = openpyxl.load_workbook("10_05_24_Test/data/loginErrors.xlsx")
        sheet = excelFile["LoginErrors"]
        rows = sheet.max_row
        data = []
        for i in range(2, rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            expected_error = sheet.cell(i,3).value
            username = username if username is not None else ""
            password = password if password is not None else ""
            expected_error = expected_error if expected_error is not None else ""
            data.append((username,password,expected_error))
        return data
    
    @pytest.mark.parametrize(("username", "password", "expected_error"), getExcelData())
    def test_login_errors(self, username, password, expected_error):
        self.login(username, password)
        error_message = self.driver.find_element(By.CSS_SELECTOR, "#login_button_container > div > form > div.error-message-container.error")
        assert error_message.text == expected_error

    # -Kullanıcı adı "standard_user" şifre "secret_sauce" gönderildiğinde kullanıcı "/inventory.html" sayfasına gönderilmelidir. Giriş yapıldıktan sonra kullanıcıya gösterilen ürün sayısı "6" adet olmalıdır.
    def test_success_login(self):
        self.login("standard_user", "secret_sauce")
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#inventory_container > div')))
        find6InventoryItemContainer = self.driver.find_element(By.CSS_SELECTOR, '#inventory_container > div')
        find6InventoryItem = find6InventoryItemContainer.find_elements(By.CLASS_NAME, "inventory_item")
        assert len(find6InventoryItem) == 6
        
    # Successful logout will be tested
    def test_success_logout(self):
        self.login("standard_user", "secret_sauce")
        WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID, "react-burger-menu-btn")))
        menu_button = self.driver.find_element(By.ID, "react-burger-menu-btn")
        menu_button.click()
        WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.ID, "logout_sidebar_link")))
        logout_link = self.driver.find_element(By.ID, "logout_sidebar_link")
        logout_link.click()
        WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located((By.ID, sauceDemoConsts.login_button_id)))
        assert self.driver.current_url == sauceDemoConsts.BASE_URL
    
    #---------------New Cases---------------#
    # See if the Price(low to high) works properly
    def test_ascending_price_filter(self):
        self.login("standard_user", "secret_sauce")
        WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, "product_sort_container")))
        filterBtn = self.driver.find_element(By.CLASS_NAME, "product_sort_container")
        filterBtn.click()
        sort_select = Select(filterBtn)
        sort_select.select_by_value("lohi")
        prices_after_sorting = []
        price_elements = self.driver.find_elements(By.CLASS_NAME, "inventory_item_price")
        for element in price_elements:
            price_txt = element.text
            price_value = float(price_txt.strip("$"))
            prices_after_sorting.append(price_value)
        for i in range(len(prices_after_sorting) - 1):
            assert prices_after_sorting[i] <= prices_after_sorting[i + 1], "Fiyatlar küçükten büyüğe sıralanmamış."
    
    # Test the equality between json file items data and website data
    def getInventoryJsonData():
        with open("10_05_24_Test/data/inventory_data.json", "r") as file:
            data = json.load(file)
        return  data["inventory_items"]
    
    @pytest.mark.parametrize(("inventory_items"), getInventoryJsonData())
    def test_inventory_details(self, inventory_items):
        self.login("standard_user", "secret_sauce")
        inventory_details = []
        items = WebDriverWait(self.driver, 10).until(EC.visibility_of_all_elements_located((By.CLASS_NAME, "inventory_item")))
        for item in items:
            name_element = WebDriverWait(item, 10).until(EC.visibility_of_element_located((By.CLASS_NAME, "inventory_item_name")))
            description_element = item.find_element(By.CLASS_NAME, "inventory_item_desc")
            name = name_element.text.strip()
            description = description_element.text.strip()
            inventory_details.append({"name": name, "description": description})
        assert inventory_items in inventory_details
    
    # Add the first item to cart and checkout 
    def test_add_and_checkout_item(self):
        self.login("standard_user", "secret_sauce")
        addBtn = WebDriverWait(self.driver,5).until(EC.element_to_be_clickable((By.ID, "add-to-cart-sauce-labs-backpack")))
        addBtn.click()
        WebDriverWait(self.driver,5).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#shopping_cart_container > a > span")))
        cartBtn = self.driver.find_element(By.CLASS_NAME, "shopping_cart_link")
        cartBtn.click()
        checkoutBtn = WebDriverWait(self.driver,5).until(EC.visibility_of_element_located((By.ID, "checkout")))
        checkoutBtn.click()
        fN = WebDriverWait(self.driver,5).until(EC.visibility_of_element_located((By.ID,"first-name")))
        fN.click()
        fN.send_keys("Miray")
        lN = self.driver.find_element(By.ID, "last-name")
        lN.click()
        lN.send_keys("Sönmez")
        pC = self.driver.find_element(By.ID, "postal-code")
        pC.click()
        pC.send_keys("00000")
        continueBtn = self.driver.find_element(By.ID, "continue")
        continueBtn.click()
        finishBtn = WebDriverWait(self.driver,5).until(EC.element_to_be_clickable((By.ID, "finish")))
        finishBtn.click()
        assert WebDriverWait(self.driver,5).until(EC.visibility_of_element_located((By.ID, "checkout_complete_container")))
    
