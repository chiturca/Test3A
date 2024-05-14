import pytest
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from constants import globalConstants

class Test_SauceDemo:
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(globalConstants.BASE_URL)

    def teardown_method(self):
        self.driver.quit()

    def getData():
        excelFile = openpyxl.load_workbook("class/10_05_24/data/invalidLogin.xlsx")
        sheet = excelFile["Sheet1"]
        rows = sheet.max_row
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))
        return data

    @pytest.mark.parametrize("username,password",getData())
    def test_invalid_login(self,username,password):
        usernameInput = self.driver.find_element(By.ID,globalConstants.username_id)
        usernameInput.send_keys(username)
        passwordInput = self.driver.find_element(By.NAME,globalConstants.password_id)
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID,globalConstants.login_button_id)
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"

    def test_valid_login(self):
        usernameInput = self.driver.find_element(By.ID,globalConstants.username_id)
        usernameInput.send_keys("standard_user")
        passwordInput = self.driver.find_element(By.NAME,globalConstants.password_id)
        passwordInput.send_keys("secret_sauce")
        loginButton = self.driver.find_element(By.ID,globalConstants.login_button_id)
        loginButton.click()
        appLogo = self.driver.find_element(By.CLASS_NAME,"app_logo")
        assert appLogo.text == "Swag Labs"